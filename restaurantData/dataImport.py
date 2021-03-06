# import of restaurant data
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import calendar

# Load workbook and print sheet names
wb = load_workbook('Restaurant_DB.xlsx')
sheet_list = wb.sheetnames
print(wb.sheetnames)

# import of tables
restaurants = pd.read_excel('Restaurant_DB.xlsx', sheet_name="Restaurants")
dishes = pd.read_excel('Restaurant_DB.xlsx', sheet_name="Gerichte")
drinks = pd.read_excel('Restaurant_DB.xlsx', sheet_name="Getränke")
particularities = pd.read_excel('Restaurant_DB.xlsx', sheet_name="Besonderheiten")
allergens = pd.read_excel('Restaurant_DB.xlsx', sheet_name="Allergene")

# get current time, needed for enquiry if restaurant open
now = datetime.now()
current_time = now.strftime("%H:%M:%S")
print("Current Time is :", current_time)
print(datetime.today().weekday())

# get weekday, needed for enquiry if restaurant open
curr_date = datetime.today()
print(calendar.day_name[curr_date.weekday()])

# needed to display all columns
pd.set_option('display.max_columns', 20)
pd.set_option('display.width', 2000)

# example enquiries
# TODO: function that select all restaurants which are open right now
print(restaurants[restaurants["Öffnet_Abends"] != "durchgehend geöffnet"])
# display one column of one restaurant
print(restaurants[restaurants["Restaurant_Name"] == "Namaste"].Beschreibung)
