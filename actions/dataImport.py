# import of restaurant data
# from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

from iteration_utilities import deepflatten

# TODO: "wie möchtest du genannt werden?" mit einbauen am Anfang damit Freitext vorhanden ist
# TODO: DB nochmal in GitHub ziehen, damit Typo in Bezeichnung korrigiert wird(Unterkategorien)
# import for sorting list
import operator

# import calendar

# Load workbook and print sheet names
# wb = load_workbook('https://github.com/schmcklr/rasa_chatbot_master/blob/kevin/restaurantData/DB_Restaurant.xlsx')
# sheet_list = wb.sheetnames
# print(wb.sheetnames)

github_link = 'https://github.com/schmcklr/rasa_chatbot_master/blob/master/restaurantData/Restaurant_DB.xlsx?raw=true'

# import of tables
restaurants = pd.read_excel(
    github_link,
    sheet_name="Restaurants")
dishes = pd.read_excel(
    github_link,
    sheet_name="Gerichte")
drinks = pd.read_excel(
    github_link,
    sheet_name="Getränke")
particularities = pd.read_excel(
    github_link,
    sheet_name="Besonderheiten")
allergens = pd.read_excel(
    github_link,
    sheet_name="Allergene")
categories = pd.read_excel(
    github_link,
    sheet_name="Speiserichtungen")
ingredients = pd.read_excel(
    github_link,
    sheet_name="Zutaten")
sub_cats = pd.read_excel(
    github_link,
    sheet_name="Unterkategorien"
)


# get current time, needed for enquiry if restaurant open
# now = datetime.now()
# current_time = now.strftime("%H:%M:%S")
# print("Current Time is :", current_time)
# print(datetime.today().weekday())

# get weekday, needed for enquiry if restaurant open
# curr_date = datetime.today()
# print(calendar.day_name[curr_date.weekday()])

# needed to display all columns
pd.set_option('display.max_columns', 20)
pd.set_option('display.width', 2000)

# example enquiries
# TODO: function that select all restaurants which are open right now
# print(restaurants[restaurants["Öffnet_Abends"] != "durchgehend geöffnet"])
# display one column of one restaurant
# print(restaurants[restaurants["Restaurant_Name"] == "Namaste"].Beschreibung)

# print(ingredients.Kategorie)
# print(sub_cats)

# slots for testing
Orientierung = "vegan"
Essensrichtung = ['bayerisch', 'mexikanisch', 'spanisch']
Protein_Liste = ['Fleisch', 'Tofu']
Carbs_Liste = ['Kartoffel']
Green_Liste = ['Gemüse', 'Sesam']

# write all dishes in a python list
dishes_list = dishes.values.tolist()

# split subcategories in different str and delete empty space
for i in range(len(dishes_list)):
    sub_category_list = dishes_list[i][11].split(",")
    scl = [x.strip() for x in sub_category_list]
    dishes_list[i][11] = scl
# print(dishes_list)

# add points to dish which fit to the information from slots - food direction
for i in range(len(dishes_list)):
    for j in range(len(dishes_list[i])):
        for elem in range(len(Essensrichtung)):
            if Essensrichtung[elem] == dishes_list[i][j]:
                dishes_list[i].insert(len(dishes_list), 2.0)
# print(dishes_list)

# add points for proteins on top of last points

for i in range(len(dishes_list)):
    for prot in range(len(Protein_Liste)):
        if Protein_Liste[prot] in (dishes_list[i][11]):
            # check if there is already a value for ranking the dish- if yes add score to this value,
            # if not append new score on last position
            if type(dishes_list[i][-1]) == float:
                dishes_list[i][-1] += 1.5
            else:
                dishes_list[i].insert(len(dishes_list), 1.5)
# print(dishes_list)

# add points for carbs on top of last points (in case of no ranking, append a float on the end of the list)
for i in range(len(dishes_list)):
    for carb in range(len(Carbs_Liste)):
        if Carbs_Liste[carb] in (dishes_list[i][11]):
            if type(dishes_list[i][-1]) == float:
                dishes_list[i][-1] += 1.2
            else:
                dishes_list[i].insert(len(dishes_list), 1.2)
# print(dishes_list)

# add points for carbs on top of last points (in case of no ranking, append a float on the end of the list)
for i in range(len(dishes_list)):
    for green in range(len(Green_Liste)):
        if Green_Liste[green] in (dishes_list[i][11]):
            if type(dishes_list[i][-1]) == float:
                dishes_list[i][-1] += 1.0
            else:
                dishes_list[i].insert(len(dishes_list), 1.0)
# print(dishes_list)

# filtering the dish_list regarding to users choice about food orientation with list comprehension
filtered_dish_list = []
for i in range(len(dishes_list)):
    if Orientierung == "vegan":
        filtered_dish_list = [x for x in dishes_list if 'vegan' in x]
    elif Orientierung == "vegetarian":
        filtered_dish_list = [x for x in dishes_list if 'vegetarian' or 'vegan' in x]
    elif Orientierung == "eat_all":
        filtered_dish_list = [x for x in dishes_list if 'eat_all' or 'vegan' or 'vegetarian' in x]

sorted_dish_list = []
for i in range(len(filtered_dish_list)):
    if type(filtered_dish_list[i][-1]) == float:
        sorted_dish_list.append(filtered_dish_list[i])
        sorted_dish_list.sort(key=lambda x: x[-1], reverse=True)

for i in range(len(sorted_dish_list)):
    print(sorted_dish_list[i])



# for i in range(len(dishes_list)-1):
#   if (range(0, 14)):
#      if dishes_list[i + 1][13] > dishes_list[i][13]:
#           dishes_list[i], dishes_list[i + 1] = dishes_list[i + 1], dishes_list[i]
# print(dishes_list)

# return_dishes name für Rückgabe
# if type(dishes_list[i][-1]) == float
